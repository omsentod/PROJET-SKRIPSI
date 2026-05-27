#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
distribusi_harga.py
================================================================================
Menganalisis, merangkum, dan memvisualisasikan kelengkapan data harga serta
distribusi statistik dari 3 sumber data utama: Wisata, Hotel, dan Tempat Makan.

Output:
1. Dashboard visualisasi yang menakjubkan di terminal (dengan tabel Unicode & ANSI Colors).
2. File Excel premium hasil analisis ('hasil-htm/distribusi_harga_summary.xlsx')
   dengan 5 sheet rapih, formatting rupiah (Rp), zebra-striping, auto-width,
   dan Grafik Batang Dinamis (Native Excel Chart).
================================================================================
"""

import os
import sys
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ==============================================================================
# KONFIGURASI FILE & PATH FALLBACK
# ==============================================================================
OUTPUT_DIR = 'hasil-htm'
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, 'distribusi_harga_summary.xlsx')

DATASETS_CONFIG = {
    'Wisata': {
        'filename': 'wisataV2_updated.xlsx',
        'display_name': 'Tempat Wisata',
        'color': '\033[92m'  # Light Green
    },
    'Hotel': {
        'filename': 'hotelV2-htm.xlsx',
        'display_name': 'Akomodasi (Hotel)',
        'color': '\033[96m'  # Light Cyan
    },
    'Tempat Makan': {
        'filename': 'tempat_makanV2-htm.xlsx',
        'display_name': 'Kuliner (Tempat Makan)',
        'color': '\033[93m'  # Light Yellow
    }
}

# ANSI Escape Colors untuk Visualisasi Terminal
C_RESET = '\033[0m'
C_BOLD = '\033[1m'
C_RED = '\033[91m'
C_GREEN = '\033[92m'
C_YELLOW = '\033[93m'
C_BLUE = '\033[94m'
C_MAGENTA = '\033[95m'
C_CYAN = '\033[96m'

# ==============================================================================
# UTILITY FUNCTIONS
# ==============================================================================
def find_file(filename):
    """Mencari file di beberapa direktori alternatif secara aman."""
    search_paths = [
        os.path.join('hasil-htm', filename),
        filename,
        os.path.join('SCRAPING', 'hasil-htm', filename),
        os.path.join('..', 'hasil-htm', filename),
        os.path.join('..', filename),
        os.path.join('data new', filename),
        os.path.join('..', 'data new', filename)
    ]
    for path in search_paths:
        if os.path.exists(path):
            return path
    return None

def is_empty_price(val):
    """Mengecek apakah nilai estimasi harga kosong (0, NaN, null)."""
    if pd.isna(val):
        return True
    try:
        f_val = float(val)
        return f_val <= 0
    except:
        return True

def parse_wisata_sumber(domain):
    """Mengkategorikan domain sumber data wisata untuk kebutuhan akademik."""
    if pd.isna(domain) or not isinstance(domain, str):
        return 'N/A'
    d = domain.strip().lower()
    if d == 'n/a' or d == '':
        return 'N/A'
    elif 'traveloka' in d:
        return 'Traveloka'
    elif 'tiket.com' in d:
        return 'Tiket.com'
    elif 'instagram' in d:
        return 'Instagram'
    elif 'google' in d:
        return 'Google Search/Snippet'
    elif 'travelspromo' in d:
        return 'Travelspromo'
    elif any(x in d for x in ['kompas.com', 'tribunnews.com', 'liputan6.com', 'detik.com', 'tempo.co', 'jawapos.com', 'kumparan.com', 'idntimes.com']):
        return 'Portal Berita (Kompas/Tribun/dll)'
    else:
        return 'Lainnya (Web/Blog Wisata)'

def format_rupiah(val):
    """Format angka menjadi string rupiah."""
    if pd.isna(val):
        return "Rp 0"
    return f"Rp {int(val):,}".replace(',', '.')

def hitung_bracket(harga, kategori):
    """Mengkategorikan harga ke dalam bracket / rentang tertentu."""
    if is_empty_price(harga):
        return 'Gratis / Kosong'
    
    h = float(harga)
    if h == 0:
        return 'Gratis (Rp 0)'
    elif h < 25000:
        return 'Sangat Murah (< Rp 25k)'
    elif h < 100000:
        return 'Murah (Rp 25k - Rp 100k)'
    elif h < 300000:
        return 'Sedang (Rp 100k - Rp 300k)'
    elif h < 1000000:
        return 'Mahal (Rp 300k - Rp 1M)'
    else:
        return 'Sangat Mahal (> Rp 1M)'

# ==============================================================================
# MAIN EXECUTION
# ==============================================================================
def main():
    print(f"{C_BOLD}{C_BLUE}="*80)
    print(f"📊 SISTEM ANALISIS DISTRIBUSI & KELENGKAPAN HARGA DATASET SKRIPSI")
    print(f"="*80 + C_RESET)
    
    # 1. Pemuatan Data
    dfs = {}
    for key, config in DATASETS_CONFIG.items():
        path = find_file(config['filename'])
        if path:
            try:
                dfs[key] = pd.read_excel(path)
                print(f"✓ {C_GREEN}Berhasil memuat{C_RESET} {config['display_name']} ({len(dfs[key])} baris) -> {path}")
            except Exception as e:
                print(f"✗ {C_RED}Gagal membaca {config['filename']}: {e}{C_RESET}")
        else:
            print(f"✗ {C_RED}File '{config['filename']}' tidak ditemukan di lokasi manapun!{C_RESET}")

    if len(dfs) < 3:
        print(f"\n{C_RED}[⚠️ WARNING] Beberapa file data tidak lengkap. Silakan pastikan file-file xlsx berada di hasil-htm/{C_RESET}")
        print("Mencoba membuat dataset simulasi (mock) agar program tetap bisa berjalan...")
        
        # Mocking data fallback
        if 'Wisata' not in dfs:
            dfs['Wisata'] = pd.DataFrame({
                'Id_Tempat': range(1, 101),
                'Nama_Tempat': [f'Wisata Alam Mock {i}' for i in range(1, 101)],
                'Estimasi_Harga': [0 if i % 10 == 0 else 15000 + (i%5)*5000 for i in range(1, 101)],
                'Sumber_Data': ['Google Snippet' if i%3==0 else 'travelspromo.com' if i%3==1 else 'traveloka.com' for i in range(1, 101)]
            })
            print("✓ Mock Wisata dibuat (100 baris)")
        if 'Hotel' not in dfs:
            dfs['Hotel'] = pd.DataFrame({
                'Id_Tempat': range(1, 81),
                'Nama_Tempat': [f'Hotel Indah Mock {i}' for i in range(1, 81)],
                'Estimasi_Harga': [250000 + (i%8)*50000 for i in range(1, 81)],
                'Sumber_Data': ['Traveloka' if i%5!=0 else 'Tiket.com' for i in range(1, 81)]
            })
            print("✓ Mock Hotel dibuat (80 baris)")
        if 'Tempat Makan' not in dfs:
            dfs['Tempat Makan'] = pd.DataFrame({
                'Id_Tempat': range(1, 121),
                'Nama_Tempat': [f'Warung Lezat Mock {i}' for i in range(1, 121)],
                'Estimasi_Harga': [15000 + (i%6)*5000 for i in range(1, 121)],
                'Sumber_Data': ['GoFood' for i in range(1, 121)]
            })
            print("✓ Mock Tempat Makan dibuat (120 baris)")

    print(f"\n{C_BOLD}{C_CYAN}1. ANALISIS KELENGKAPAN DATA HARGA (ADA VS KOSONG){C_RESET}")
    
    # Menghitung Kelengkapan
    kelengkapan_rows = []
    total_gab_total = 0
    total_gab_ada = 0
    total_gab_kosong = 0
    
    daftar_kosong = [] # Untuk menyimpan rincian item kosong
    
    for key, df in dfs.items():
        total_rows = len(df)
        col_harga = 'Estimasi_Harga'
        
        # Hitung ada vs kosong
        kosong_mask = df[col_harga].apply(is_empty_price)
        kosong_count = kosong_mask.sum()
        ada_count = total_rows - kosong_count
        
        pct_ada = (ada_count / total_rows) * 100
        pct_kosong = (kosong_count / total_rows) * 100
        
        kelengkapan_rows.append({
            'Kategori': DATASETS_CONFIG[key]['display_name'],
            'Total Data': total_rows,
            'Ada Harga': ada_count,
            'Persen Ada': pct_ada,
            'Harga Kosong': kosong_count,
            'Persen Kosong': pct_kosong
        })
        
        total_gab_total += total_rows
        total_gab_ada += ada_count
        total_gab_kosong += kosong_count
        
        # Kumpulkan data kosong untuk log detail
        df_kosong = df[kosong_mask].copy()
        for _, r in df_kosong.iterrows():
            daftar_kosong.append({
                'Kategori': DATASETS_CONFIG[key]['display_name'],
                'Id Tempat': r.get('Id_Tempat', '-'),
                'Nama Tempat': r.get('Nama_Tempat', 'Tanpa Nama'),
                'Sumber Sebelum': r.get('Sumber_Data', 'N/A')
            })

    # Cetak Tabel Kelengkapan di Terminal
    print("┌──────────────────────────┬────────────┬─────────────────────┬─────────────────────┬─────────────────┐")
    print("│ Kategori                 │ Total Data │ Ada Harga           │ Harga Kosong        │ Kelengkapan (%) │")
    print("├──────────────────────────┼────────────┼─────────────────────┼─────────────────────┼─────────────────┤")
    for r in kelengkapan_rows:
        display_kat = r['Kategori'][:24].ljust(24)
        total_d = str(r['Total Data']).rjust(10)
        ada_h = f"{r['Ada Harga']} ({r['Persen Ada']:.1f}%)".rjust(19)
        kosong_h = f"{r['Harga Kosong']} ({r['Persen Kosong']:.1f}%)".rjust(19)
        kelengkapan_pct = f"{r['Persen Ada']:.1f}%".rjust(15)
        print(f"│ {display_kat} │ {total_d} │ {ada_h} │ {kosong_h} │ {kelengkapan_pct} │")
    
    # Baris Total Gabungan
    pct_gab_ada = (total_gab_ada / total_gab_total) * 100
    pct_gab_kosong = (total_gab_kosong / total_gab_total) * 100
    print("├──────────────────────────┼────────────┼─────────────────────┼─────────────────────┼─────────────────┤")
    print(f"│ {C_BOLD}Total Gabungan           {C_RESET}│ {C_BOLD}{str(total_gab_total).rjust(10)}{C_RESET} │ {C_BOLD}{f'{total_gab_ada} ({pct_gab_ada:.1f}%)'.rjust(19)}{C_RESET} │ {C_RED}{f'{total_gab_kosong} ({pct_gab_kosong:.1f}%)'.rjust(19)}{C_RESET} │ {C_BOLD}{C_GREEN}{f'{pct_gab_ada:.1f}%'.rjust(15)}{C_RESET} │")
    print("└──────────────────────────┴────────────┴─────────────────────┴─────────────────────┴─────────────────┘")

    # 2. Analisis Statistik Deskriptif (hanya data berharga > 0)
    print(f"\n{C_BOLD}{C_CYAN}2. ANALISIS STATISTIK DESKRIPTIF DATA HARGA (Rp){C_RESET}")
    statistik_rows = []
    
    for key, df in dfs.items():
        col_harga = 'Estimasi_Harga'
        valid_prices = df[~df[col_harga].apply(is_empty_price)][col_harga].astype(float)
        
        if len(valid_prices) > 0:
            s_min = valid_prices.min()
            s_max = valid_prices.max()
            s_mean = valid_prices.mean()
            s_median = valid_prices.median()
            s_std = valid_prices.std()
        else:
            s_min = s_max = s_mean = s_median = s_std = 0
            
        statistik_rows.append({
            'Kategori': DATASETS_CONFIG[key]['display_name'],
            'Min': s_min,
            'Max': s_max,
            'Mean': s_mean,
            'Median': s_median,
            'Std Dev': s_std
        })

    # Cetak Tabel Statistik di Terminal
    print("┌──────────────────────────┬─────────────────┬─────────────────┬─────────────────┬─────────────────┬─────────────────┐")
    print("│ Kategori                 │ Min Harga       │ Max Harga       │ Rata-rata (Mean)│ Median          │ Standar Deviasi │")
    print("├──────────────────────────┼─────────────────┼─────────────────┼─────────────────┼─────────────────┼─────────────────┤")
    for r in statistik_rows:
        display_kat = r['Kategori'][:24].ljust(24)
        h_min = format_rupiah(r['Min']).rjust(15)
        h_max = format_rupiah(r['Max']).rjust(15)
        h_mean = format_rupiah(r['Mean']).rjust(15)
        h_median = format_rupiah(r['Median']).rjust(15)
        h_std = format_rupiah(r['Std Dev']).rjust(15)
        print(f"│ {display_kat} │ {h_min} │ {h_max} │ {h_mean} │ {h_median} │ {h_std} │")
    print("└──────────────────────────┴─────────────────┴─────────────────┴─────────────────┴─────────────────┴─────────────────┘")

    # 3. Analisis Distribusi Portal/Platform Sumber Data
    print(f"\n{C_BOLD}{C_CYAN}3. ANALISIS DISTRIBUSI PORTAL / PLATFORM ASAL SUMBER DATA{C_RESET}")
    sumber_distribusi = []
    
    for key, df in dfs.items():
        total_rows = len(df)
        
        # Standarisasi kolom Sumber Data
        if 'Sumber_Data' in df.columns:
            if key == 'Wisata':
                df['Platform_Clean'] = df['Sumber_Data'].apply(parse_wisata_sumber)
            else:
                df['Platform_Clean'] = df['Sumber_Data'].fillna('N/A').apply(lambda x: str(x).strip())
        else:
            df['Platform_Clean'] = 'N/A'
            
        counts = df['Platform_Clean'].value_counts()
        for plat, count in counts.items():
            pct = (count / total_rows) * 100
            sumber_distribusi.append({
                'Kategori': DATASETS_CONFIG[key]['display_name'],
                'Platform': plat,
                'Jumlah': count,
                'Persentase': pct
            })

    # Urutkan berdasarkan Kategori, lalu Jumlah terbesar
    sumber_distribusi_df = pd.DataFrame(sumber_distribusi)
    
    # Cetak Tabel Distribusi Portal di Terminal
    print("┌──────────────────────────┬────────────────────────────────────────┬─────────────┬─────────────┐")
    print("│ Kategori                 │ Portal / Platform Sumber Data          │ Jumlah Data │ Kontribusi  │")
    print("├──────────────────────────┼────────────────────────────────────────┼─────────────┼─────────────┤")
    last_kat = ""
    for idx, r in sumber_distribusi_df.iterrows():
        # Bersihkan cetakan kategori agar tidak duplikat vertikal
        if r['Kategori'] == last_kat:
            display_kat = "".ljust(24)
        else:
            display_kat = r['Kategori'][:24].ljust(24)
            last_kat = r['Kategori']
            if idx > 0:
                print("├──────────────────────────┼────────────────────────────────────────┼─────────────┼─────────────┤")
                
        display_plat = r['Platform'][:38].ljust(38)
        jumlah = str(r['Jumlah']).rjust(11)
        persen = f"{r['Persentase']:.1f}%".rjust(11)
        print(f"│ {display_kat} │ {display_plat} │ {jumlah} │ {persen} │")
    print("└──────────────────────────┴────────────────────────────────────────┴─────────────┴─────────────┘")

    # 4. Analisis Distribusi Rentang Harga (Price Brackets)
    print(f"\n{C_BOLD}{C_CYAN}4. ANALISIS DISTRIBUSI RENTANG HARGA (PRICE BRACKETS){C_RESET}")
    bracket_distribusi = []
    
    # Definisi urutan bracket agar rapi
    bracket_order = [
        'Gratis / Kosong', 'Gratis (Rp 0)', 'Sangat Murah (< Rp 25k)', 
        'Murah (Rp 25k - Rp 100k)', 'Sedang (Rp 100k - Rp 300k)', 
        'Mahal (Rp 300k - Rp 1M)', 'Sangat Mahal (> Rp 1M)'
    ]
    
    for key, df in dfs.items():
        total_rows = len(df)
        df['Price_Bracket'] = df['Estimasi_Harga'].apply(lambda x: hitung_bracket(x, key))
        
        counts = df['Price_Bracket'].value_counts()
        
        # Masukkan dalam urutan yang konsisten
        for b_name in bracket_order:
            if b_name in counts:
                count = counts[b_name]
                pct = (count / total_rows) * 100
                bracket_distribusi.append({
                    'Kategori': DATASETS_CONFIG[key]['display_name'],
                    'Rentang': b_name,
                    'Jumlah': count,
                    'Persentase': pct
                })

    bracket_distribusi_df = pd.DataFrame(bracket_distribusi)
    
    # Cetak Tabel Bracket Harga di Terminal
    print("┌──────────────────────────┬────────────────────────────────────────┬─────────────┬─────────────┐")
    print("│ Kategori                 │ Rentang Harga (Bracket)                │ Jumlah Data │ Persentase  │")
    print("├──────────────────────────┼────────────────────────────────────────┼─────────────┼─────────────┤")
    last_kat = ""
    for idx, r in bracket_distribusi_df.iterrows():
        if r['Kategori'] == last_kat:
            display_kat = "".ljust(24)
        else:
            display_kat = r['Kategori'][:24].ljust(24)
            last_kat = r['Kategori']
            if idx > 0:
                print("├──────────────────────────┼────────────────────────────────────────┼─────────────┼─────────────┤")
                
        display_rentang = r['Rentang'][:38].ljust(38)
        jumlah = str(r['Jumlah']).rjust(11)
        persen = f"{r['Persentase']:.1f}%".rjust(11)
        print(f"│ {display_kat} │ {display_rentang} │ {jumlah} │ {persen} │")
    print("└──────────────────────────┴────────────────────────────────────────┴─────────────┴─────────────┘")

    # Informasikan Item Kosong yang Perlu Diperbaiki
    if daftar_kosong:
        print(f"\n{C_BOLD}{C_YELLOW}⚠️  DITEMUKAN {len(daftar_kosong)} DATA HARGA KOSONG:{C_RESET}")
        print("Tabel 5 teratas (Daftar lengkap akan diexport ke file Excel):")
        print("┌──────────────────────────┬───────────┬──────────────────────────────────────────┬────────────────────┐")
        print("│ Kategori                 │ ID Tempat │ Nama Tempat                              │ Sumber Pengganti   │")
        print("├──────────────────────────┼───────────┼──────────────────────────────────────────┼────────────────────┤")
        for i, item in enumerate(daftar_kosong[:5]):
            display_kat = item['Kategori'][:24].ljust(24)
            display_id = str(item['Id Tempat']).rjust(9)
            display_nama = item['Nama Tempat'][:40].ljust(40)
            display_sumber = str(item['Sumber Sebelum'])[:18].ljust(18)
            print(f"│ {display_kat} │ {display_id} │ {display_nama} │ {display_sumber} │")
        print("└──────────────────────────┴───────────┴──────────────────────────────────────────┴────────────────────┘")
        if len(daftar_kosong) > 5:
            print(f"  ... dan {len(daftar_kosong)-5} data kosong lainnya.")

    # ==============================================================================
    # EXPORT KE EXCEL DENGAN STYLING PREMIUM & GRAFIK DYNAMIC NATIVE
    # ==============================================================================
    print(f"\n{C_BOLD}{C_CYAN}💾 MENYUSUN WORKBOOK EXCEL PREMIUM...{C_RESET}")
    
    # 1. Konversi baris data ke dataframe siap cetak
    df_kelengkapan_ex = pd.DataFrame(kelengkapan_rows)
    df_statistik_ex = pd.DataFrame(statistik_rows)
    
    # Tambahkan baris total gabungan untuk Sheet Kelengkapan di Excel
    total_row_df = pd.DataFrame([{
        'Kategori': 'Total Gabungan',
        'Total Data': total_gab_total,
        'Ada Harga': total_gab_ada,
        'Persen Ada': pct_gab_ada,
        'Harga Kosong': total_gab_kosong,
        'Persen Kosong': pct_gab_kosong
    }])
    df_kelengkapan_ex = pd.concat([df_kelengkapan_ex, total_row_df], ignore_index=True)
    
    df_daftar_kosong_ex = pd.DataFrame(daftar_kosong) if daftar_kosong else pd.DataFrame(columns=['Kategori', 'Id Tempat', 'Nama Tempat', 'Sumber Sebelum'])

    # Style definitions
    font_family = "Segoe UI"
    navy_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    accent_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=11)
    font_total = Font(name=font_family, size=11, bold=True, color="000000")
    font_title = Font(name=font_family, size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name=font_family, size=10, italic=True, color="595959")
    
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom_side = Side(border_style="double", color="1F4E79")
    thin_top_side = Side(border_style="thin", color="1F4E79")
    total_border = Border(left=thin_border_side, right=thin_border_side, top=thin_top_side, bottom=double_bottom_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Create Excel Writer
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        # Tulis dataframe mentah dulu, baru kita modifikasi styles-nya
        df_kelengkapan_ex.to_excel(writer, sheet_name='Ringkasan Kelengkapan', index=False)
        sumber_distribusi_df.to_excel(writer, sheet_name='Distribusi Portal Sumber', index=False)
        df_statistik_ex.to_excel(writer, sheet_name='Statistik Deskriptif', index=False)
        bracket_distribusi_df.to_excel(writer, sheet_name='Distribusi Rentang Harga', index=False)
        df_daftar_kosong_ex.to_excel(writer, sheet_name='Daftar Harga Kosong', index=False)

    # Membuka kembali file menggunakan openpyxl untuk merapikan visual & styling
    wb = openpyxl.load_workbook(OUTPUT_EXCEL)
    
    # --------------------------------------------------------------------------
    # SHEET 1: RINGKASAN KELENGKAPAN (DENGAN NATIVE CHART)
    # --------------------------------------------------------------------------
    ws = wb['Ringkasan Kelengkapan']
    ws.insert_rows(1, 4) # Insert 4 blank rows at the top for title block
    
    # Tulis Banner Judul
    ws['A1'] = "LAPORAN RINGKASAN KELENGKAPAN DATA HARGA"
    ws['A1'].font = font_title
    ws['A2'] = "Mengukur data 'Ada Harga' vs 'Harga Kosong' (Penelitian Skripsi Malang Raya)"
    ws['A2'].font = font_subtitle
    
    # Styling Table Headers
    header_row = 5
    for col_idx in range(1, 7):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
    
    # Formatting Data Rows
    total_data_rows = len(df_kelengkapan_ex)
    for r_idx in range(6, 6 + total_data_rows):
        is_total_row = (r_idx == 5 + total_data_rows)
        row_fill = total_fill if is_total_row else (gray_fill if r_idx % 2 == 0 else PatternFill(fill_type=None))
        row_font = font_total if is_total_row else font_data
        row_border = total_border if is_total_row else thin_border
        
        # Kategori
        c_kat = ws.cell(row=r_idx, column=1)
        c_kat.fill = row_fill
        c_kat.font = row_font
        c_kat.alignment = align_left
        c_kat.border = row_border
        
        # Total Data
        c_tot = ws.cell(row=r_idx, column=2)
        c_tot.fill = row_fill
        c_tot.font = row_font
        c_tot.alignment = align_right
        c_tot.border = row_border
        c_tot.number_format = '#,##0'
        
        # Ada Harga
        c_ada = ws.cell(row=r_idx, column=3)
        c_ada.fill = row_fill
        c_ada.font = row_font
        c_ada.alignment = align_right
        c_ada.border = row_border
        c_ada.number_format = '#,##0'
        
        # Persen Ada
        c_pada = ws.cell(row=r_idx, column=4)
        c_pada.fill = row_fill
        c_pada.font = row_font
        c_pada.alignment = align_right
        c_pada.border = row_border
        # Convert percent value from [0-100] to standard excel fraction [0-1]
        c_pada.value = float(c_pada.value) / 100
        c_pada.number_format = '0.0%'
        
        # Harga Kosong
        c_kos = ws.cell(row=r_idx, column=5)
        c_kos.fill = row_fill
        c_kos.font = row_font
        c_kos.alignment = align_right
        c_kos.border = row_border
        c_kos.number_format = '#,##0'
        
        # Persen Kosong
        c_pkos = ws.cell(row=r_idx, column=6)
        c_pkos.fill = row_fill
        c_pkos.font = row_font
        c_pkos.alignment = align_right
        c_pkos.border = row_border
        c_pkos.value = float(c_pkos.value) / 100
        c_pkos.number_format = '0.0%'
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
    ws.column_dimensions['A'].width = 28
    
    # Pastikan gridlines terlihat
    ws.views.sheetView[0].showGridLines = True
    
    # NATIVE EXCEL BAR CHART
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Analisis Kelengkapan Data: Ada Harga vs Harga Kosong"
    chart.y_axis.title = "Jumlah Data"
    chart.x_axis.title = "Kategori"
    
    # Mengambil referensi data (Ada Harga di kolom C, Harga Kosong di kolom E)
    # Kita buat chart hanya untuk 3 kategori utama (baris 6 sampai 8), lewati total gabungan
    data_ref = Reference(ws, min_col=3, min_row=5, max_col=3, max_row=8)
    data_ref_kosong = Reference(ws, min_col=5, min_row=5, max_col=5, max_row=8)
    cats_ref = Reference(ws, min_col=1, min_row=6, max_row=8)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.add_data(data_ref_kosong, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    # Custom warna chart (Series 1: Blue, Series 2: Red)
    chart.height = 14
    chart.width = 22
    
    # Letakkan chart di bawah tabel (Baris 11, Kolom A)
    ws.add_chart(chart, "A11")
    
    # --------------------------------------------------------------------------
    # SHEET 2: DISTRIBUSI PORTAL SUMBER
    # --------------------------------------------------------------------------
    ws_portal = wb['Distribusi Portal Sumber']
    ws_portal.insert_rows(1, 4)
    ws_portal['A1'] = "DISTRIBUSI DATA BERDASARKAN PORTAL ASAL"
    ws_portal['A1'].font = font_title
    ws_portal['A2'] = "Detail kontribusi masing-masing platform portal (Instagram, GoFood, Traveloka, dsb)"
    ws_portal['A2'].font = font_subtitle
    
    # Format headers
    for col_idx in range(1, 5):
        cell = ws_portal.cell(row=5, column=col_idx)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Format data
    total_portal_rows = len(sumber_distribusi_df)
    for r_idx in range(6, 6 + total_portal_rows):
        fill = gray_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        c_kat = ws_portal.cell(row=r_idx, column=1)
        c_kat.fill = fill
        c_kat.font = font_data
        c_kat.alignment = align_left
        c_kat.border = thin_border
        
        c_plat = ws_portal.cell(row=r_idx, column=2)
        c_plat.fill = fill
        c_plat.font = font_data
        c_plat.alignment = align_left
        c_plat.border = thin_border
        
        c_qty = ws_portal.cell(row=r_idx, column=3)
        c_qty.fill = fill
        c_qty.font = font_data
        c_qty.alignment = align_right
        c_qty.border = thin_border
        c_qty.number_format = '#,##0'
        
        c_pct = ws_portal.cell(row=r_idx, column=4)
        c_pct.fill = fill
        c_pct.font = font_data
        c_pct.alignment = align_right
        c_pct.border = thin_border
        c_pct.value = float(c_pct.value) / 100
        c_pct.number_format = '0.0%'
        
    for col in ws_portal.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_portal.column_dimensions[col_letter].width = max(max_len + 3, 14)
    ws_portal.column_dimensions['A'].width = 28
    ws_portal.column_dimensions['B'].width = 38
    ws_portal.views.sheetView[0].showGridLines = True
    
    # --------------------------------------------------------------------------
    # SHEET 3: STATISTIK DESKRIPTIF
    # --------------------------------------------------------------------------
    ws_stat = wb['Statistik Deskriptif']
    ws_stat.insert_rows(1, 4)
    ws_stat['A1'] = "STATISTIK DESKRIPTIF DATA HARGA VALID"
    ws_stat['A1'].font = font_title
    ws_stat['A2'] = "Kalkulasi nilai min, max, median, rata-rata, dan standar deviasi (harga kosong diabaikan)"
    ws_stat['A2'].font = font_subtitle
    
    for col_idx in range(1, 7):
        cell = ws_stat.cell(row=5, column=col_idx)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(6, 6 + len(df_statistik_ex)):
        fill = gray_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        # Kategori
        ws_stat.cell(row=r_idx, column=1).alignment = align_left
        
        # Harga columns
        for col_idx in range(2, 7):
            c_val = ws_stat.cell(row=r_idx, column=col_idx)
            c_val.alignment = align_right
            c_val.number_format = 'Rp#,##0'
            
        for col_idx in range(1, 7):
            c = ws_stat.cell(row=r_idx, column=col_idx)
            c.fill = fill
            c.font = font_data
            c.border = thin_border
            
    for col in ws_stat.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_stat.column_dimensions[col_letter].width = max(max_len + 3, 16)
    ws_stat.column_dimensions['A'].width = 28
    ws_stat.views.sheetView[0].showGridLines = True
    
    # --------------------------------------------------------------------------
    # SHEET 4: DISTRIBUSI RENTANG HARGA (BRACKETS)
    # --------------------------------------------------------------------------
    ws_brack = wb['Distribusi Rentang Harga']
    ws_brack.insert_rows(1, 4)
    ws_brack['A1'] = "ANALISIS BRACKET / RENTANG HARGA AKADEMIS"
    ws_brack['A1'].font = font_title
    ws_brack['A2'] = "Pengelompokan data ke dalam rentang harga yang disesuaikan untuk skripsi"
    ws_brack['A2'].font = font_subtitle
    
    for col_idx in range(1, 5):
        cell = ws_brack.cell(row=5, column=col_idx)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(6, 6 + len(bracket_distribusi_df)):
        fill = gray_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        c_kat = ws_brack.cell(row=r_idx, column=1)
        c_kat.fill = fill
        c_kat.font = font_data
        c_kat.alignment = align_left
        c_kat.border = thin_border
        
        c_rent = ws_brack.cell(row=r_idx, column=2)
        c_rent.fill = fill
        c_rent.font = font_data
        c_rent.alignment = align_left
        c_rent.border = thin_border
        
        c_qty = ws_brack.cell(row=r_idx, column=3)
        c_qty.fill = fill
        c_qty.font = font_data
        c_qty.alignment = align_right
        c_qty.border = thin_border
        c_qty.number_format = '#,##0'
        
        c_pct = ws_brack.cell(row=r_idx, column=4)
        c_pct.fill = fill
        c_pct.font = font_data
        c_pct.alignment = align_right
        c_pct.border = thin_border
        c_pct.value = float(c_pct.value) / 100
        c_pct.number_format = '0.0%'
        
    for col in ws_brack.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_brack.column_dimensions[col_letter].width = max(max_len + 3, 14)
    ws_brack.column_dimensions['A'].width = 28
    ws_brack.column_dimensions['B'].width = 38
    ws_brack.views.sheetView[0].showGridLines = True
    
    # --------------------------------------------------------------------------
    # SHEET 5: DAFTAR HARGA KOSONG
    # --------------------------------------------------------------------------
    ws_empty = wb['Daftar Harga Kosong']
    ws_empty.insert_rows(1, 4)
    ws_empty['A1'] = "DAFTAR DATA HARGA KOSONG (AUDIT TRAIL)"
    ws_empty['A1'].font = font_title
    ws_empty['A2'] = "Daftar tempat wisata/hotel/kuliner yang harga tiket/kamarnya masih bernilai Rp 0 atau NaN"
    ws_empty['A2'].font = font_subtitle
    
    for col_idx in range(1, 5):
        cell = ws_empty.cell(row=5, column=col_idx)
        cell.fill = PatternFill(start_color="A51D24", end_color="A51D24", fill_type="solid") # Dark Red Header for Warning
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(6, 6 + len(df_daftar_kosong_ex)):
        fill = accent_fill if r_idx % 2 == 0 else PatternFill(fill_type=None)
        
        c_kat = ws_empty.cell(row=r_idx, column=1)
        c_kat.fill = fill
        c_kat.font = font_data
        c_kat.alignment = align_left
        c_kat.border = thin_border
        
        c_id = ws_empty.cell(row=r_idx, column=2)
        c_id.fill = fill
        c_id.font = font_data
        c_id.alignment = align_center
        c_id.border = thin_border
        
        c_nam = ws_empty.cell(row=r_idx, column=3)
        c_nam.fill = fill
        c_nam.font = font_data
        c_nam.alignment = align_left
        c_nam.border = thin_border
        
        c_src = ws_empty.cell(row=r_idx, column=4)
        c_src.fill = fill
        c_src.font = font_data
        c_src.alignment = align_left
        c_src.border = thin_border
        
    for col in ws_empty.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_empty.column_dimensions[col_letter].width = max(max_len + 3, 14)
    ws_empty.column_dimensions['A'].width = 28
    ws_empty.column_dimensions['C'].width = 45
    ws_empty.views.sheetView[0].showGridLines = True
    
    # Save styled workbook
    wb.save(OUTPUT_EXCEL)
    print(f"✓ {C_GREEN}Excel Premium Sukses Disimpan{C_RESET} -> {OUTPUT_EXCEL}")
    print(f"{C_BOLD}{C_BLUE}="*80)
    print(f"🎉 ANALISIS DISTRIBUSI SELESAI DENGAN SEMPURNA!")
    print(f"Silakan buka file excel hasil rekap untuk lampiran Skripsi Anda.")
    print(f"="*80 + C_RESET)

if __name__ == "__main__":
    main()
